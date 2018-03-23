import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# questa funzione ricerca i dati su internet.
# nel dettaglio data una lista di url di prodotti su amazon restituisce una lista con nome e prezzo per ogni url
def scrape(urls_list):
    data_list = []
    for i, url in enumerate(urls_list):
        #richiesta http
        r = requests.get(url)
        # parso codice html della pagina
        soup = BeautifulSoup(r.text, "html.parser")
        try:
            # ricavo nome prodotto
            # questi "finders" sono stati ottenuti ispezionando il codice html
            nomeProdotto = soup.find("span", {"id": "productTitle"})
            # pulisco la stringa ottenuta
            nomeProdotto_text = " ".join(nomeProdotto.text.split())
        except:
            # se non ho trovato il nome del prodotto, l'url non è di un prodotto di amazon.
            data_list = data_list + [("url non valido", "url non valido")]
            print("Warning: url numero", i+1, "non valido")
            continue #passo al prossimo url
        #ricavo prezzo prodotto
        try:
            # ho trovato tre casi di identificazione per i prezzi dei prodotti
            prezzo = soup.find("span", {"id": "priceblock_dealprice"})
            prezzo_text = prezzo.text
        except:
            try:
                prezzo = soup.find("span", {"id": "priceblock_ourprice"})
                prezzo_text = prezzo.text
            except:
                try:
                    prezzo = soup.find("span", {"class": "a-color-price"})
                    prezzo_text = prezzo.text
                except:
                    data_list = data_list + [(nomeProdotto_text, "Prodotto non disponibile")]
                    continue  # passo al prossimo url
        # aggiungo i dati ricavati in coda alla lista
        data_list = data_list + [(nomeProdotto_text, prezzo_text)]
    return data_list


# questa funzione legge degli url da una colonna di un file excel
def get_urls(file_name, column):
    # carico file
    wb = load_workbook(file_name)
    # carico foglio di lavoro attivo
    ws = wb.active
    # recupero colonna desiderata
    col = ws[column]
    # costruisco e ritorno una lista con gli url
    return [cell.value for cell in col]


# questa funzione salva due colonne di dati su un file excel
def save_data(xlsx_file, data_list, col_1, col_2):
    # carico file
    wb = load_workbook(xlsx_file)
    # carico foglio di lavoro attivo
    ws = wb.active
    # ciclo sulla lista di coppie da inserire
    for i, data_tuple in enumerate(data_list):  # ['a','b','c'] --enumerate--> [(0,'a'),(1,'b'),(2,'c')]
        # inserisco dati nel file excel, gli indici delle righe su excel partono da uno
        ws[col_1+str(i+1)] = data_tuple[0]
        ws[col_2+str(i+1)] = data_tuple[1]
    # sovrascrivo il file
    wb.save(xlsx_file)

# assegno il nome del file a una variabile
file = "data.xlsx"
# recupero gli url
urls = get_urls(file, "A")
# recupero i dati dagli url trovati
data = scrape(urls)
# salvo i dati trovati
save_data(file, data, "B", "C")
