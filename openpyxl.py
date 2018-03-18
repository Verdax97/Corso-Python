# openpyxl

from openpyxl import Workbook

# creo un istanza della classe Woorkbook
wb = Workbook()
# Prendo worksheet attivo
ws = wb.active

# assegno 42 alla cella A1
ws['A1'] = 42
# inserisco una riga in fondo
ws.append([1, 2, 3])

wb.save("sample.xlsx")



from openpyxl import load_workbook

# carico woorkbook
wb = load_workbook('sample.xlsx')
# stampo i nomi di tutti i worksheet
print(wb.sheetnames)
# prendo il worksheet attivo
ws = wb.active
# stampo il valore della cella A1
print(ws["A1"].value)
# assegno due valori
ws['A10'] = 7
ws['B8'] = 6
# inserisco una riga
ws.append([3, 4, 5])
wb.save('sample.xlsx')
